from asyncio import run
from datetime import date
from pathlib import Path
from typing import Optional

from httpx import AsyncHTTPTransport
from openpyxl import Workbook

from bestdori import get_songs_all, get_bands_all, get_song
from bestdori.model import Song, BandsAll, DifficultyInt, Language

TRANSPORT = AsyncHTTPTransport(retries=3)  # 设置重试次数
TIMEOUT = 30  # 设置超时时间（秒）
SAVEPATH = Path(__file__).parent / 'bestdori_charts.xlsx'  # 设置保存路径

EXCLUDE_SONG_ID = [
    273,  # [原曲] final phase
    1000,  # Legendary (English Version)
    1001,  # Excellent (Hey, Let's Go!) (English Version)
    1004,  # See the world
    10001,  # 彩虹节拍
]


class ExcelRow(object):

    def __init__(self, song_id: int, song: Song, bands_all: BandsAll, difficulty: DifficultyInt):
        self.song_id = song_id
        self.song = song
        self.bands_all = bands_all.__root__
        self.difficulty = difficulty

    @property
    def chart_id(self) -> int:
        """返回谱面ID，区别于歌曲ID，若谱面难度为Special，则谱面ID为歌曲ID+10000，否则等于歌曲ID"""
        return self.song_id + 10000 if self.difficulty == DifficultyInt.Special else self.song_id

    @property
    def band(self) -> str:
        """返回该歌曲的乐队名"""
        return self.bands_all[self.song.bandId].bandName[Language.Japanese]

    @property
    def tag(self) -> str:
        """返回该歌曲的分类"""
        match self.song.tag:
            case 'normal':
                return 'Original'  # 原创曲目
            case 'anime':
                return 'Cover'  # 翻唱曲目
            case 'tie_up':
                return 'Extra'  # 联动/合作/原唱曲目
        return 'Unknown'

    @property
    def bpm(self) -> str:
        """返回该歌曲的BPM或BPM区间"""
        bpm_list = [bpm.bpm for bpm in self.song.bpm[self.difficulty]]
        bpm_max = max(bpm_list)
        bpm_min = min(bpm_list)
        if bpm_max == bpm_min:
            return str(bpm_max)
        else:
            return f'{bpm_min}-{bpm_max}'

    @property
    def release_jp(self) -> Optional[date]:
        """返回该歌曲在日服的发布日期"""
        return self.get_release(Language.Japanese)

    @property
    def release_cn(self) -> Optional[date]:
        """返回该歌曲在国服的发布日期"""
        return self.get_release(Language.ChineseSimplified)

    @property
    def note(self) -> int:
        """返回该谱面的总物量"""
        return self.song.notes[self.difficulty]

    @property
    def duration(self) -> str:
        """返回该谱面的总时长，格式为1:23.4"""
        duration = self.song.length
        return f'{int(duration // 60)}:{int(duration % 60):02d}.{int(duration * 10 % 10):0d}'

    @property
    def title(self) -> str:
        """返回该歌曲的标题"""
        return self.song.musicTitle[Language.Japanese]

    @property
    def level(self) -> int:
        """返回该谱面的等级，例如25、26"""
        return self.song.difficulty[self.difficulty].playLevel

    @property
    def level_class(self) -> str:
        """返回该谱面的等级分类，例如Expert、Special"""
        return self.difficulty.name

    @property
    def lyricist(self) -> str:
        """返回该歌曲的作词者"""
        return self.song.lyricist[Language.Japanese]

    @property
    def composer(self) -> str:
        """返回该歌曲的作曲者"""
        return self.song.composer[Language.Japanese]

    @property
    def arranger(self) -> str:
        """返回该歌曲的编曲者"""
        return self.song.arranger[Language.Japanese]

    def get_release(self, server: Language):
        if self.difficulty == DifficultyInt.Special:
            if (published_at := self.song.difficulty[DifficultyInt.Special].publishedAt) is not None:
                publish_datetime = published_at[server]
            else:
                publish_datetime = self.song.publishedAt[server]
        else:
            publish_datetime = self.song.publishedAt[server]

        if publish_datetime is None:
            return None
        return publish_datetime.date()


async def main():
    headline = [
        'Id',
        'Band',
        'Tag',
        'BPM',
        'Release(JP)',
        'Release(CN)',
        'Note',
        'Duration',
        'Title',
        'Level',
        'Class',
        'Lyricist',
        'Composer',
        'Arranger',
    ]
    # 对应BandoriChartInfo类中的属性名，将使用__getattribute__来获取属性
    column_to_write = {
        1: 'chart_id',
        2: 'band',
        3: 'tag',
        4: 'bpm',
        5: 'release_jp',
        6: 'release_cn',
        7: 'note',
        8: 'duration',
        9: 'title',
        10: 'level',
        11: 'level_class',
        12: 'lyricist',
        13: 'composer',
        14: 'arranger',
    }

    book = Workbook()
    sheet_all = book.create_sheet(title='charts_all', index=0)
    sheet_ex = book.create_sheet(title='charts_ex', index=1)
    sheet_sp = book.create_sheet(title='charts_sp', index=2)

    # 写标题行
    for col, item in enumerate(headline, start=1):
        sheet_all.cell(column=col, row=1, value=item)
        sheet_ex.cell(column=col, row=1, value=item)
        sheet_sp.cell(column=col, row=1, value=item)

    songs_all = await get_songs_all(transport=TRANSPORT, timeout=TIMEOUT)
    bands_all = await get_bands_all(transport=TRANSPORT, timeout=TIMEOUT)

    all_chart_list: list[ExcelRow] = []
    ex_chart_list: list[ExcelRow] = []
    sp_chart_list: list[ExcelRow] = []

    for song_id, song_meta in songs_all.__root__.items():
        if song_id in EXCLUDE_SONG_ID:
            continue
        song = await get_song(song_id, transport=TRANSPORT, timeout=TIMEOUT)
        print(f'正在处理 {song_id} {song.musicTitle[Language.Japanese]}')

        # 处理 ex 谱面
        ex_row = ExcelRow(song_id, song, bands_all, DifficultyInt.Expert)
        ex_chart_list.append(ex_row)
        all_chart_list.append(ex_row)
        print(f'  EX: {ex_row.bpm} {ex_row.note} {ex_row.level} {ex_row.level_class}')

        # 处理 sp 谱面
        if len(song.difficulty) == 5:
            sp_row = ExcelRow(song_id, song, bands_all, DifficultyInt.Special)
            sp_chart_list.append(sp_row)
            all_chart_list.append(sp_row)
            print(f'  SP: {sp_row.bpm} {sp_row.note} {sp_row.level} {sp_row.level_class}')

    # 把all、ex、sp列表分别写入三张表，注意row是行，col是列
    for sheet, chart_list in [(sheet_ex, ex_chart_list), (sheet_sp, sp_chart_list)]:
        for chart in chart_list:
            for col, key in column_to_write.items():
                sheet.cell(column=col, row=chart.song_id + 1, value=chart.__getattribute__(key))

    for row, chart in enumerate(all_chart_list, start=2):
        for col, key in column_to_write.items():
            sheet_all.cell(column=col, row=row, value=chart.__getattribute__(key))

    book.save(SAVEPATH)


if __name__ == '__main__':
    run(main())
